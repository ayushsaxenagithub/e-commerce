from django.contrib import admin
from .models import Category, Product


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'description')


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'category', 'price', 'quantity')
    list_filter = ('category',)
    search_fields = ('name', 'description')
    filter_horizontal = ('photos', 'videos', 'tags')
    actions = ['download_selected_as_excel']

    def download_selected_as_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        from django.http import HttpResponse

        # Create a new workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Set headers
        headers = ['Name', 'Category', 'Description', 'Price', 'Quantity']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = header

        # Write data rows
        for row_num, product in enumerate(queryset, 2):
            sheet[f'A{row_num}'] = product.name
            sheet[f'B{row_num}'] = str(product.category)
            sheet[f'C{row_num}'] = product.description
            sheet[f'D{row_num}'] = product.price
            sheet[f'E{row_num}'] = product.quantity

        # Save the workbook to a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=products.xlsx'
        workbook.save(response)

        return response

    download_selected_as_excel.short_description = 'Download selected as Excel'


# Register other models and customize their admins if needed
